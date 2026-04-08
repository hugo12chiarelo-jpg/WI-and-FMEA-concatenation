import glob
import os
import sys
import json
import time
import pandas as pd
from openai import OpenAI, APIConnectionError, APIStatusError, RateLimitError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---
API_KEY_DS = os.environ.get("API_KEY_DS")

if not API_KEY_DS:
    print("Error: API_KEY_DS environment variable not set.")
    sys.exit(1)

client = OpenAI(api_key=API_KEY_DS, base_url="https://api.deepseek.com", timeout=120.0)

# File paths
# Input files must be placed in the data/ folder of this repository:
#   - Work Instructions: data/work instruction/*.xlsx  (all .xlsx files in the folder)
#   - FMEA:              any single .xlsx file in the data/ folder (not inside a subdirectory)
WI_DIR = "data/work instruction"
DATA_DIR = "data"
OUTPUT_PATH = "results/analysis.xlsx"

# FMEA spreadsheet column names that hold the maintainable item and failure mechanism
FMEA_COL_ITEM = "Unnamed: 6"
FMEA_COL_MECHANISM = "Unnamed: 7"
# FMEA left-table columns: component names and symptom codes
FMEA_COL_LEFT_LABEL = "Unnamed: 3"
FMEA_COL_LEFT_GROUP = "Unnamed: 0"

# Coverage colour fills
FILL_COVERED = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FILL_PARTIAL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_NOT_COVERED = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")

# Header colour fill (teal, matching the example image)
FILL_HEADER = PatternFill(start_color="1F6B75", end_color="1F6B75", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# --- Locate FMEA file dynamically ---
def find_fmea_file():
    """Return the path to the single FMEA .xlsx file located directly inside DATA_DIR."""
    candidates = [
        f for f in glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]
    if not candidates:
        print(f"Error: no .xlsx FMEA file found in '{DATA_DIR}'")
        sys.exit(1)
    if len(candidates) > 1:
        print(f"Warning: multiple .xlsx files found in '{DATA_DIR}', using: {candidates[0]}")
    return candidates[0]


# --- Load data ---
def load_data():
    try:
        wi_files = [
            f for f in glob.glob(os.path.join(WI_DIR, "*.xlsx"))
            if not os.path.basename(f).startswith("~$")
        ]
        if not wi_files:
            print(f"Error loading data: no .xlsx files found in '{WI_DIR}'")
            sys.exit(1)

        fmea_path = find_fmea_file()
        df_fmea = pd.read_excel(fmea_path)
        print(
            f"✓ Data loaded: FMEA '{os.path.basename(fmea_path)}', "
            f"{len(wi_files)} Work Instruction file(s)."
        )
        return wi_files, df_fmea
    except SystemExit:
        raise
    except Exception as e:
        print(f"Error loading data: {e}")
        sys.exit(1)


# --- Prepare WI questions with unique IDs ---
def prepare_wi_questions(wi_files):
    """Load all WI files and return a list of dicts, each with a unique question_id.

    ID format (in priority order):
      1. If an 'Unnamed: 4' column exists and contains a value like 'PM00248A-Q4', use it directly.
      2. Otherwise construct '{WTT_ID}-{Work instruction task}', e.g. 'PM00009A-2.d'.
         The WTT_ID is taken from the 'WTT ID' column (forward-filled) or, when absent/NaN,
         from the filename stem (e.g. 'PM00072A.xlsx' → 'PM00072A').
    """
    questions = []
    for filepath in wi_files:
        df = pd.read_excel(filepath)
        filename_wtt = os.path.splitext(os.path.basename(filepath))[0]

        # Forward-fill WTT ID column so merged cells propagate
        if "WTT ID" in df.columns:
            df["WTT ID"] = df["WTT ID"].ffill()

        for _, row in df.iterrows():
            # --- Determine question ID ---
            q_id = None

            # Priority 1: pre-built ID in Unnamed: 4 (e.g. PM00248A-Q4)
            if "Unnamed: 4" in df.columns:
                val = row.get("Unnamed: 4")
                if pd.notna(val):
                    candidate = str(val).strip()
                    if candidate and candidate.lower() != "nan":
                        q_id = candidate

            # Priority 2: construct from WTT ID + task number
            if not q_id:
                raw_wtt = row.get("WTT ID", None) if "WTT ID" in df.columns else None
                wtt = (
                    str(raw_wtt).strip()
                    if pd.notna(raw_wtt) and str(raw_wtt).strip().lower() != "nan"
                    else filename_wtt
                )
                task = row.get("Work instruction task", None)
                task_str = (
                    str(task).strip()
                    if pd.notna(task) and str(task).strip().lower() != "nan"
                    else ""
                )
                q_id = f"{wtt}-{task_str}" if task_str else wtt

            # --- Extract other fields ---
            maintainable_item = str(row.get("Maintainable Item", "")).strip()
            reporting_question = str(row.get("Reporting Question", "")).strip()

            # Skip rows with no actual reporting question
            if not reporting_question or reporting_question.lower() == "nan":
                continue

            questions.append(
                {
                    "question_id": q_id,
                    "maintainable_item": maintainable_item,
                    "reporting_question": reporting_question,
                }
            )

    return questions


# --- Extract FMEA component → symptoms mapping ---
def extract_fmea_component_symptoms(df_fmea):
    """Parse the FMEA left-side pivot table and return a dict mapping
    each component name to its list of symptom/failure-mode codes."""
    # Identify all known component names from the group column
    known_components = set(
        str(v).strip()
        for v in df_fmea[FMEA_COL_LEFT_GROUP].dropna().unique()
        if str(v).strip() not in ("Row Labels", "Grand Total")
    )

    component_symptoms = {}
    current_comp = None
    for _, row in df_fmea.iterrows():
        label_val = row.get(FMEA_COL_LEFT_LABEL)
        if pd.isna(label_val):
            continue
        label = str(label_val).strip()
        if not label or label in ("Row Labels", "Grand Total"):
            continue
        if label in known_components:
            current_comp = label
            component_symptoms.setdefault(current_comp, [])
        elif current_comp and label not in known_components:
            component_symptoms[current_comp].append(label)

    return component_symptoms


# --- Extract FMEA component/mechanism pairs ---
def extract_fmea_rows(df_fmea, component_symptoms=None):
    """Parse the FMEA spreadsheet and return a list of dicts with
    'maintainable_item', 'failure_mechanism', and 'symptoms', skipping header/total rows."""
    if component_symptoms is None:
        component_symptoms = {}
    rows = []
    current_item = None
    for _, row in df_fmea.iterrows():
        item_val = row.get(FMEA_COL_ITEM)
        mech_val = row.get(FMEA_COL_MECHANISM)

        # Skip pure header row
        if str(item_val).strip() == "Maintainable Item":
            continue

        # A non-null item cell starts a new component group (ignore "* Total" rows)
        if pd.notna(item_val):
            label = str(item_val).strip()
            if "Total" in label or label == "Grand Total":
                current_item = None
                continue
            current_item = label

        # A mechanism cell with a valid current component is a data row
        if current_item and pd.notna(mech_val):
            mech_str = str(mech_val).strip()
            if mech_str == "Failure Mechanism":
                continue
            symptoms = component_symptoms.get(current_item, [])
            rows.append(
                {
                    "maintainable_item": current_item,
                    "failure_mechanism": mech_str,
                    "symptoms": ", ".join(symptoms) if symptoms else "Unknown",
                }
            )
    return rows


# --- Build prompt for deepseek-reasoner ---
def build_prompt(wi_questions, fmea_rows):
    """Build the analysis prompt using a structured WI question list (with IDs) and FMEA rows."""

    # Build a readable table of WI questions with their IDs
    wi_lines = []
    for q in wi_questions:
        wi_lines.append(
            f"  [{q['question_id']}] ({q['maintainable_item']}) {q['reporting_question']}"
        )
    wi_text = "\n".join(wi_lines)

    # Build FMEA list with symptoms context
    fmea_lines = "\n".join(
        f"- {r['maintainable_item']} | {r['failure_mechanism']} "
        f"| Symptoms: {r['symptoms']}"
        for r in fmea_rows
    )

    prompt = f"""
You are a maintenance engineering expert specialized in rotating equipment and ISO 14224 failure classification.
You apply Reliability-Centered Maintenance (RCM) principles to evaluate maintenance strategy coverage.

You have two datasets:

1. **Work Instructions – Reporting Questions:**
Each line follows the format: [QUESTION_ID] (Maintainable Item) Reporting Question text
```
{wi_text}
```

2. **FMEA – Failure Mechanisms to analyse (with observed symptoms per component):**
Each line follows the format: Component | Failure Mechanism | Symptoms (ISO 14224 codes)
```
{fmea_lines}
```

---

**TASK:**
For EVERY failure mechanism listed above, determine whether it is covered by any reporting question
in the Work Instructions. You MUST return one entry per mechanism — do not skip any.

---

**COVERAGE CLASSIFICATION RULES:**

- **"Covered"**: A reporting question directly or functionally identifies, detects, or prevents
  that failure mechanism. This includes:
  - A question that explicitly mentions the physical phenomenon (e.g., "corrosion" → covers
    mechanism "2.2 Corrosion").
  - A question linked to a component that is functionally responsible for the mechanism.
  - A question that monitors a symptom that is a direct and specific consequence of the mechanism.

- **"Partial"**: A question may indicate the mechanism indirectly or unreliably, for example:
  - A generic symptom question (e.g., "unusual vibration") that could relate to multiple
    mechanisms (misalignment, imbalance, looseness, wear, etc.).
  - A question about a parent or nearby component rather than the exact failing component.
  - The mechanism is only one of several causes that could trigger the observed symptom.

- **"Not covered"**: Absolutely no reporting question, even indirectly, can detect or prevent
  the mechanism. Use this only when there is truly no applicable question.

---

**ENGINEERING KNOWLEDGE — apply these associations:**
- Vibration (VIB) questions → partially cover: misalignment, imbalance, looseness, wear,
  bearing clearance failure, shaft deflection.
- Noise (NOI) questions → partially cover: wear, erosion, looseness, bearing failure, cavitation.
- Temperature / overheating (OHE) questions → partially cover: lack of lubrication, friction,
  overheating, bearing failure.
- Lube oil analysis / sampling → cover or partially cover: wear, fatigue, contamination,
  corrosion, erosion in lubricated components.
- Visual inspection for leaks/corrosion/damage → cover: external leakage (ELP/ELU), corrosion,
  deformation, structural deficiency.
- Seal gas / buffer gas parameter questions → cover or partially cover: seal-related leakage,
  dry gas seal failure mechanisms.
- Instrument/calibration/signal questions → cover: signal/indication failures, spurious alarms,
  control failures, out-of-adjustment mechanisms.
- Differential pressure questions on filters/strainers → cover: blockage/plugged mechanisms.

---

**FUNCTIONAL MAPPING RULES:**
- If a question is associated with a parent component (e.g., "*Centrifugal Compressor"), it
  applies to all sub-components (casing, bearings, impeller, shaft, diffuser, coupling) unless
  explicitly excluded.
- Questions about instrumentation (transmitters, panels, indicators, alarms, calibration) apply
  to components that depend on control/signal, such as dry gas seal, anti-surge valve, lube oil
  pump, seal gas panel.
- A question about the lube oil system applies to all lubricated components (bearings, shaft,
  coupling, etc.).
- Use the Symptom codes (ISO 14224) listed for each component to associate symptom-detecting
  questions with the corresponding failure mechanisms. A question that detects a symptom listed
  for a component covers or partially covers all mechanisms that can produce that symptom.

---

**CRITERIA:**
- Do NOT classify as "Covered" if the relation is very remote or speculative.
- Prefer "Partial" when there is reasonable doubt.
- Use "Not covered" only when absolutely no question addresses the mechanism.

---

**OUTPUT FORMAT:**
Return a single valid JSON object (no markdown fences, no extra text) with:

1. Top-level key **"analysis"**: an array where each element has exactly:
   - "maintainable_item": component name from the FMEA list
   - "failure_mechanism": failure mechanism from the FMEA list
   - "coverage": one of "Covered", "Partial", "Not covered"
   - "note": brief justification in the same language as the Work Instructions data
   - "question_ids": list of QUESTION_IDs (from the bracketed IDs in the Work Instructions above,
     e.g. "PM00248A-Q4", "PM00009A-2.f") that support the coverage.
     IMPORTANT: always include the full ID including the question number
     (e.g. "PM00009A-2.f", NOT just "PM00009A").
     Use an empty list [] if "Not covered".

2. Top-level key **"gaps"**: an array of objects for mechanisms with "Not covered" coverage:
   - "maintainable_item": component name
   - "failure_mechanism": mechanism name
   - "recommendation": a brief recommendation for a new inspection task to cover this gap

3. Top-level key **"excesses"**: an array of objects where multiple questions are fully redundant
   for the same mechanism (i.e., 3 or more question_ids that all directly cover the same thing):
   - "maintainable_item": component name
   - "failure_mechanism": mechanism name
   - "redundant_question_ids": list of question IDs that are redundant

4. Top-level key **"recommendations"**: a string with a concise summary (3-5 bullet points) of
   the main recommendations to improve the maintenance strategy coverage.

Only output valid JSON, no extra text, no markdown fences.
"""
    return prompt


# --- Call DeepSeek Reasoner API ---
def call_deepseek(prompt, max_retries=3):
    """Call the DeepSeek Reasoner API with retry/backoff on transient errors.

    Note: ``deepseek-reasoner`` does NOT support the ``response_format``
    parameter, so JSON is requested via the prompt only.
    """
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model="deepseek-reasoner",
                messages=[
                    {
                        "role": "system",
                        "content": "You are a maintenance reliability expert. Output only valid JSON.",
                    },
                    {"role": "user", "content": prompt},
                ],
            )
            print("✓ DeepSeek Reasoner analysis complete.")
            return response.choices[0].message.content
        except (APIConnectionError, RateLimitError) as e:
            if attempt < max_retries:
                wait = 2 ** attempt  # exponential backoff: 2 s, 4 s, 8 s (attempts 1–3)
                print(f"⚠ DeepSeek API error (attempt {attempt}/{max_retries}): {e}. Retrying in {wait}s…")
                time.sleep(wait)
            else:
                print(f"Error calling DeepSeek API: {e}")
                sys.exit(1)
        except APIStatusError as e:
            print(f"Error calling DeepSeek API: HTTP {e.status_code} - {e.message}")
            sys.exit(1)
        except Exception as e:
            print(f"Error calling DeepSeek API: {e}")
            sys.exit(1)


# --- Save results as Excel workbook ---
def save_excel(analysis_items, gaps=None, excesses=None, recommendations=""):
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: FMEA Coverage Analysis ──────────────────────────────────────
    ws = wb.active
    ws.title = "FMEA Coverage Analysis"

    headers = [
        "Maintainable Item (FMEA)",
        "Failure Mechanism",
        "Coverage",
        "Note",
        "Reporting Question ID identification",
    ]

    # Write and style header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, item in enumerate(analysis_items, start=2):
        coverage = item.get("coverage", "")
        question_ids = item.get("question_ids", [])
        ids_str = ", ".join(str(q) for q in question_ids) if question_ids else ""

        values = [
            item.get("maintainable_item", ""),
            item.get("failure_mechanism", ""),
            coverage,
            item.get("note", ""),
            ids_str,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        # Apply coverage colour to the Coverage cell (column C)
        cov_cell = ws.cell(row=row_idx, column=3)
        if coverage == "Covered":
            cov_cell.fill = FILL_COVERED
            cov_cell.font = Font(bold=True, color="FFFFFF")
        elif coverage == "Partial":
            cov_cell.fill = FILL_PARTIAL
            cov_cell.font = Font(bold=True, color="000000")
        elif coverage == "Not covered":
            cov_cell.fill = FILL_NOT_COVERED
            cov_cell.font = Font(bold=True, color="FFFFFF")

        cov_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths: A=Maintainable Item, B=Failure Mechanism, C=Coverage,
    #                    D=Note, E=Reporting Question ID identification
    col_widths = [30, 35, 15, 60, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Auto-filter on all columns
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Gaps ─────────────────────────────────────────────────────────
    ws_gaps = wb.create_sheet(title="Gaps (Not Covered)")
    gap_headers = ["Maintainable Item", "Failure Mechanism", "Recommendation"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws_gaps.cell(row=1, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws_gaps.row_dimensions[1].height = 30

    for row_idx, gap in enumerate(gaps or [], start=2):
        for col_idx, key in enumerate(
            ["maintainable_item", "failure_mechanism", "recommendation"], start=1
        ):
            cell = ws_gaps.cell(row=row_idx, column=col_idx, value=gap.get(key, ""))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    for col_idx, width in enumerate([30, 35, 70], start=1):
        ws_gaps.column_dimensions[get_column_letter(col_idx)].width = width
    ws_gaps.freeze_panes = "A2"
    if gaps:
        ws_gaps.auto_filter.ref = ws_gaps.dimensions

    # ── Sheet 3: Excesses ─────────────────────────────────────────────────────
    ws_exc = wb.create_sheet(title="Excesses (Redundant)")
    exc_headers = ["Maintainable Item", "Failure Mechanism", "Redundant Question IDs"]
    for col_idx, header in enumerate(exc_headers, start=1):
        cell = ws_exc.cell(row=1, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws_exc.row_dimensions[1].height = 30

    for row_idx, exc in enumerate(excesses or [], start=2):
        redundant = exc.get("redundant_question_ids", [])
        values = [
            exc.get("maintainable_item", ""),
            exc.get("failure_mechanism", ""),
            ", ".join(str(q) for q in redundant),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_exc.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    for col_idx, width in enumerate([30, 35, 60], start=1):
        ws_exc.column_dimensions[get_column_letter(col_idx)].width = width
    ws_exc.freeze_panes = "A2"
    if excesses:
        ws_exc.auto_filter.ref = ws_exc.dimensions

    # ── Sheet 4: Recommendations ──────────────────────────────────────────────
    ws_rec = wb.create_sheet(title="Recommendations")
    rec_cell = ws_rec.cell(row=1, column=1, value="Strategic Recommendations")
    rec_cell.fill = FILL_HEADER
    rec_cell.font = FONT_HEADER
    rec_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_rec.column_dimensions["A"].width = 100
    ws_rec.row_dimensions[1].height = 25

    content_cell = ws_rec.cell(row=2, column=1, value=recommendations or "No recommendations provided.")
    content_cell.alignment = Alignment(vertical="top", wrap_text=True)
    # Estimate row height based on line count (each newline adds ~15 points)
    line_count = (recommendations or "").count("\n") + 1
    ws_rec.row_dimensions[2].height = max(30, min(line_count * 15, 400))

    wb.save(OUTPUT_PATH)
    print(f"✓ Results saved to {OUTPUT_PATH}")


# --- Parse API response and save ---
def process_results(json_str, fmea_rows):
    try:
        # Strip optional markdown code fences the model may add (e.g. ```json … ```)
        cleaned = json_str.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[-1]
            if cleaned.endswith("```"):
                cleaned = cleaned.rsplit("```", 1)[0]

        data = json.loads(cleaned)
        analysis = data.get("analysis", [])

        # Ensure every FMEA row is represented; add placeholder if AI missed any
        ai_keys = {
            (str(e.get("maintainable_item", "")).strip(),
             str(e.get("failure_mechanism", "")).strip())
            for e in analysis
        }
        for fmea_row in fmea_rows:
            key = (str(fmea_row["maintainable_item"]).strip(),
                   str(fmea_row["failure_mechanism"]).strip())
            if key not in ai_keys:
                analysis.append(
                    {
                        "maintainable_item": fmea_row["maintainable_item"],
                        "failure_mechanism": fmea_row["failure_mechanism"],
                        "coverage": "Not covered",
                        "note": "No analysis returned by AI for this mechanism.",
                        "question_ids": [],
                    }
                )

        covered = sum(1 for e in analysis if e.get("coverage") == "Covered")
        partial = sum(1 for e in analysis if e.get("coverage") == "Partial")
        not_covered = sum(1 for e in analysis if e.get("coverage") == "Not covered")
        print(
            f"Summary: {len(analysis)} mechanisms — "
            f"{covered} Covered, {partial} Partial, {not_covered} Not covered."
        )

        gaps = data.get("gaps", [])
        excesses = data.get("excesses", [])
        recommendations = data.get("recommendations", "")

        # Fallback: build gaps list from Not covered items if AI didn't provide it
        if not gaps:
            gaps = [
                {
                    "maintainable_item": e["maintainable_item"],
                    "failure_mechanism": e["failure_mechanism"],
                    "recommendation": "No recommendation provided.",
                }
                for e in analysis
                if e.get("coverage") == "Not covered"
            ]

        print(f"  Gaps: {len(gaps)}, Excesses: {len(excesses)}")
        save_excel(analysis, gaps=gaps, excesses=excesses, recommendations=recommendations)

    except json.JSONDecodeError:
        print("Error: API response is not valid JSON.")
        print("Response received:")
        print(json_str)
        sys.exit(1)


# --- Main ---
def main():
    print("--- FMEA Coverage Analysis with DeepSeek Reasoner ---")
    wi_files, df_fmea = load_data()
    wi_questions = prepare_wi_questions(wi_files)
    print(f"✓ Prepared {len(wi_questions)} Work Instruction questions with IDs.")
    component_symptoms = extract_fmea_component_symptoms(df_fmea)
    fmea_rows = extract_fmea_rows(df_fmea, component_symptoms)
    print(f"✓ Extracted {len(fmea_rows)} FMEA component/mechanism pairs.")
    prompt = build_prompt(wi_questions, fmea_rows)
    result_json = call_deepseek(prompt)
    process_results(result_json, fmea_rows)
    print("--- Analysis finished ---")


if __name__ == "__main__":
    main()
